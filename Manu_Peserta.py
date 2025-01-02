import openpyxl
import os

from main import menu_utama

FILE_NAME = "Data_Peserta.xlsx"
 
def peserta():
    print("\nMAU MELAKUKAN APA DIPESERTA???")
    print("1. Tambah Peserta")
    print("2. Lihat Peserta")
    print("3. Edit Peserta")
    print("4. Hapus Peserta")
    print("5. Kembali")
    pilih = input("mau yang mana??")

    if pilih =="1":
        tambah()
    elif pilih == "2":
        baca()
    elif pilih == "3":
        edit()
    elif pilih == "4":
        hapus()
    elif pilih ==  "5":
       menu_utama()
    else :
        print("salah")

def tambah():
    print("\nISI UNTUK MENAMBAHKAN NAMA PESERTA")
    Nama_Karyawan = input("Masukan Nama Karyawan: ")
    Umur = input("Masukan Umur Karyawan: ")
    telp = input("Masukan telp Karyawan: ")
    Email = input("Masukan Email Karyawan: ")
    try:
        simpan_ke_excel(Nama_Karyawan, Umur, telp, Email)
        print("Data berhasil ditambahkan ke menu peserta")
        peserta()
    except Exception as a:
        print(f"Terjadi kesalahan saat menyimpan data: {a}")


def baca():
    if not os.path.exists(FILE_NAME):
        print("DATA TIDAK DITEMUKAN")
        return
    
    workbook=openpyxl.load_workbook(FILE_NAME)
    if"Peserta" not in workbook.sheetnames:#check for the correct sheet name
        print("Sheet belum ada")
        return
    
    sheet=workbook['Peserta']#use the correct sheet name
    if sheet.max_row == 1:
        print("Belum ada data")
        return
    
    print("\nDAFTAR PESERTA")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"Nama Karyawan:{row[0]}")
        print(f"Umur:{row[1]}")
        print(f"telp:{row[2]}")
        print(f"Email:{row[3]}\n")
        workbook.close()
    peserta()
    

def edit():
    if not os.path.exists(FILE_NAME):
        print("DATA TIDAK DITEMUKAN")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if "Peserta" not in workbook.sheetnames:
        print("Sheet Belum ada")
        return

    sheet = workbook['Peserta']
    if sheet.max_row == 1:
        print("Belum ada data untuk diedit.")
        return

    # Display current entries
    print("\nDAFTAR PESERTA UNTUK DIEDIT")
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        print(f"{index}. Nama Karyawan: {row[0]}, Umur: {row[1]},telp: {row[2]}, Email: {row[3]}")

    # Ask user for the entry they want to edit
    try:
        pilihan = int(input("Masukkan nomor peserta yang ingin diedit: "))
        if pilihan < 1 or pilihan > sheet.max_row - 1:  # Adjust for header row
            print("Nomor peserta tidak valid.")
            return

        # Get the current data
        current_row = sheet[pilihan + 1]  # +1 because we skip the header row

        # Display current data
        print(f"Data saat ini: Nama Karyawan: {current_row[0].value}, Umur: {current_row[1].value},  telp: {current_row[2].value},  Email: {current_row[3].value}")

        # Get new data from the user
        new_nama_karyawan = input("Masukkan Nama Karyawan baru (tekan Enter untuk tetap): ")
        new_Umur = input("Masukkan Umur baru (tekan Enter untuk tetap): ")
        new_telp = input("Masukkan telp baru (tekan Enter untuk tetap): ")
        new_email = input("Masukkan Email baru (tekan Enter untuk tetap): ")

        # Update the row with new values
        
        if new_nama_karyawan:
            current_row[0].value = new_nama_karyawan
        if new_Umur:
            current_row[1].value = new_Umur
        if new_telp:
            current_row[2].value = new_telp
        if new_email:
            current_row[3].value = new_email

        # Save the changes
        workbook.save(FILE_NAME)
        print("Data berhasil diperbarui.")
    except Exception as e:
        print(f"Terjadi kesalahan saat mengedit data: {e}")
    finally:
        peserta()
        workbook.close()
   
                                
def hapus():
    if not os.path.exists(FILE_NAME):
        print("DATA TIDAK DITEMUKAN")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if "Peserta" not in workbook.sheetnames:
        print("Sheet Belum ada")
        return

    sheet = workbook['Peserta']
    if sheet.max_row == 1:
        print("Belum ada data untuk dihapus.")
        return

    # Display current entries
    print("\nDAFTAR KEGIATAN UNTUK DIHAPUS")
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        print(f"{index}. Nama karyawan: {row[0]}, Umur: {row[1]}, telp: {row[2]}, Email: {row[3]}")

    # Ask user for the entry they want to delete
    try:
        pilihan = int(input("Masukkan nomor peserta yang ingin dihapus: "))
        if pilihan < 1 or pilihan > sheet.max_row - 1:  # Adjust for header row
            print("Nomor peserta tidak valid.")
            return

        # Delete the selected row
        sheet.delete_rows(pilihan + 1)  # +1 because we skip the header row

        # Save the changes
        workbook.save(FILE_NAME)
        print("Data berhasil dihapus.")
    except Exception as e:
        print(f"Terjadi kesalahan saat menghapus data: {e}")
    finally:
        peserta()
        workbook.close()


def simpan_ke_excel(Nama_Karyawan, Umur, telp, email):
    """
    Menyimpan data pembelian ke file Excel.
    """
    
    # Cek jika file Excel belum ada, buat file baru
    if not os.path.exists(FILE_NAME):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Peserta"
        
        # Tambahkan header
        sheet.append(["Nama_Karyawan", "Umur", "Telp", "Email"])
    else:
        workbook = openpyxl.load_workbook(FILE_NAME)
        sheet = workbook["Peserta"]
        
    # Tambahkan data pembelian
    sheet.append([Nama_Karyawan, Umur, telp, email])
    workbook.save(FILE_NAME)
