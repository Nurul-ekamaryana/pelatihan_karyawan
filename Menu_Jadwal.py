from main import menu_utama
import openpyxl
import os

FILE_NAME = "Data_Jadwal.xlsx"

def jadwal():
    print("\nMAU MELAKUKAN APA DIJADWAL???")
    print("1. Tambah Jadwal")
    print("2. Lihat Jadwal")
    print("3. Edit Jadwal")
    print("4. Hapus Jadwal")
    print("5. Kembali")
    pilihan = input("Pilih Jadwal: ")
    
    if pilihan == "1":
        tambah()
    elif pilihan == "2":
        baca()  
    elif pilihan == "3":
        edit()
    elif pilihan == "4":
        hapus()
    elif pilihan == "5":
        menu_utama()
    else:
        print("PILIHAN SALAH!!😒😒")

def simpan_ke_excel(Nama_karyawan, Nama_Kegiatan, Email, Telp, Umur, Waktu):
    # Jika file Excel tidak ada
    if not os.path.exists(FILE_NAME):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Jadwal'
        sheet.append(['Nama_karyawan', 'Nama_Kegiatan', 'Email', 'Telp', 'Umur', 'Waktu'])
    else:
        # Jika sudah ada, buka file Excel
        workbook = openpyxl.load_workbook(FILE_NAME)
        sheet = workbook['Jadwal']

    # Tambahkan data ke Excel

    # Simpan perubahan
    sheet.append([Nama_karyawan, Nama_Kegiatan, Email, Telp, Umur, Waktu])
    workbook.save(FILE_NAME)


def baca_peserta():
    """Membaca data peserta dari file Excel dan mengembalikan daftar peserta."""
    if not os.path.exists("Data_Peserta.xlsx"):
        return []  # Jika file tidak ada, kembalikan list kosong

    workbook = openpyxl.load_workbook("Data_Peserta.xlsx")
    sheet = workbook["Peserta"]

    peserta_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        peserta_list.append(row)  # Menyimpan seluruh baris sebagai tuple

    return peserta_list

def tambah():
    print("\nISI UNTUK MENAMBAHKAN JADWAL !!")

    # Baca daftar peserta dari file peserta
    peserta_list = baca_peserta()

    # Tampilkan pilihan peserta
    print("Pilih peserta:")
    for i, peserta in enumerate(peserta_list, start=1):
        print(f"{i}. {peserta}")  # Menampilkan seluruh data peserta
    print("0. Masukkan peserta baru")

    pilihan_nama = int(input("Masukkan pilihan: "))
    if pilihan_nama == 0:
        Nama_karyawan = input("Masukkan Nama Karyawan: ")
    else:
        Nama_karyawan = peserta_list[pilihan_nama - 1][0]  
        Telp = peserta_list[pilihan_nama - 1][2]
        Umur = peserta_list[pilihan_nama - 1][1]
        Email = peserta_list[pilihan_nama - 1][3]  # Ambil email dari tuple


    # Lanjutkan dengan input lainnya...
    Nama_Kegiatan = input("Masukkan Nama Kegiatan: ")
    Waktu = input("Masukkan Waktu (format: YYYY-MM-DD HH:MM): ")

    try:
        simpan_ke_excel(Nama_karyawan, Nama_Kegiatan, Email, Telp, Umur, Waktu)
        print("Data berhasil ditambahkan ke jadwal Jadwal.")
        jadwal()
    except Exception as e:
        print(f"Terjadi kesalahan saat menyimpan data: {e}")
        
def baca():
    # input lihat jadwal
    
    if not os.path.exists(FILE_NAME):
        print("DATA TIDAK DITEMUKAN")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if "Jadwal" not in workbook.sheetnames:  # Check for the correct sheet name
        print("Sheet Belum ada")
        return
    
    sheet = workbook['Jadwal']  # Use the correct sheet name
    if sheet.max_row == 1:
        print("Belum ada data")
        return
    
    print("\nDAFTAR KEGIATAN/JADWAL")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"\nNama Karyawan: {row[0]}, ")
        print(f"Nama Kegiatan: {row[1]},")
        print(f"Email: {row[2]}")
        print(f"Telp : {row[3]}")
        print(f"Umur : {row[4]}")
        print(f"Waktu : {row[5]}\n")

    workbook.close()
    jadwal()

def edit():
    if not os.path.exists(FILE_NAME):
        print("DATA TIDAK DITEMUKAN")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if "Jadwal" not in workbook.sheetnames:
        print("Sheet Belum ada")
        return

    sheet = workbook['Jadwal']
    if sheet.max_row == 1:
        print("Belum ada data untuk diedit.")
        return

    # Display current entries
    print("\nDAFTAR KEGIATAN/JADWAL UNTUK DIEDIT")
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        print(f"{index}. Nama Karyawan: {row[0]}, Nama Kegiatan: {row[1]}, Email: {row[2]}, Telp: {row[3]}, Umur: {row[4]}, Waktu: {row[5]}")

    # Ask user for the entry they want to edit
    try:
        pilihan = int(input("Masukkan nomor jadwal yang ingin diedit: "))
        if pilihan < 1 or pilihan > sheet.max_row - 1:  # Adjust for header row
            print("Nomor jadwal tidak valid.")
            return

        # Get the current data
        current_row = sheet[pilihan + 1]  # +1 because we skip the header row

        # Display current data
        print(f"Data saat ini: Nama Karyawan: {current_row[0].value}, Nama Kegiatan: {current_row[1].value}, Email: {current_row[2].value}, Telp: {current_row[3].value}, Umur: {current_row[4].value}, Waktu: {current_row[5].value}")
 
        peserta_list = baca_peserta()

        # Tampilkan pilihan nama karyawan
        print("Pilih nama karyawan:")
        for i, peserta in enumerate(peserta_list, start=1):
            print(f"{i}. {peserta}")
        print("0. Masukkan nama karyawan baru")

        pilihan_nama = int(input("Masukkan pilihan: "))
        if pilihan_nama == 0:
            new_nama_karyawan = input("Masukkan Nama Karyawan: ")
        else:
            new_nama_karyawan = peserta_list[pilihan_nama - 1][0]
            new_tepl = peserta_list[pilihan_nama - 1][2]
            new_umur = peserta_list[pilihan_nama - 1][1]
            new_Email = peserta_list[pilihan_nama - 1][3]
            new_nama_kegiatan = input("Masukkan Nama Kegiatan baru (tekan Enter untuk tetap): ")
            new_waktu = input("Masukkan Waktu Awal baru (format: YYYY-MM-DD HH:MM, tekan Enter untuk tetap): ")

            # Update the row with new values
            if new_nama_karyawan:
                current_row[0].value = new_nama_karyawan
            if new_nama_kegiatan:
                current_row[1].value = new_nama_kegiatan
            if new_Email:
                current_row[2].value = new_Email
            if new_tepl:
                current_row[3].value = new_tepl
            if new_umur:
                current_row[4].value = new_umur
            if new_waktu:
                current_row[5].value = new_waktu

            # Save the changes
            workbook.save(FILE_NAME)
            print("Data berhasil diperbarui.")
    except Exception as e:
        print(f"Terjadi kesalahan saat mengedit data: {e}")
    finally:
        jadwal()
        workbook.close()
        
def hapus():
    if not os.path.exists(FILE_NAME):
        print("DATA TIDAK DITEMUKAN")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if "Jadwal" not in workbook.sheetnames:
        print("Sheet Belum ada")
        return

    sheet = workbook['Jadwal']
    if sheet.max_row == 1:
        print("Belum ada data untuk dihapus.")
        return

    # Display current entries
    print("\nDAFTAR JADWAL UNTUK DIHAPUS")
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        print(f"{index}. Nama Karyawan: {row[0]}, Nama Kegiatan: {row[1]}, Email: {row[2]}, Telp: {row[3]}, Umur: {row[4]}, Waktu: {row[5]}")

    # Ask user for the entry they want to delete
    try:
        pilihan = int(input("Masukkan nomor jadwal yang ingin dihapus: "))
        if pilihan < 1 or pilihan > sheet.max_row - 1:  # Adjust for header row
            print("Nomor jadwal tidak valid.")
            return

        # Delete the selected row
        sheet.delete_rows(pilihan + 1)  # +1 because we skip the header row

        # Save the changes
        workbook.save(FILE_NAME)
        print("Data berhasil dihapus.")
    except Exception as e:
        print(f"Terjadi kesalahan saat menghapus data: {e}")
    finally:
        jadwal()
        workbook.close()

