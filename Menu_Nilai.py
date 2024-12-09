import openpyxl
import os

FILE_NAME = "Data_Nilai.xlsx"

def simpan_ke_excel(Nama_karyawan, Nama_kegiatan, Waktu, Nilai):
    # Jika file Excel tidak ada
    if not os.path.exists(FILE_NAME):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Nilai'
        sheet.append(['Nama_karyawan', 'Nama_Kegiatan', 'Waktu', 'Nilai'])
    else:
        # Jika sudah ada, buka file Excel
        workbook = openpyxl.load_workbook(FILE_NAME)
        sheet = workbook['Nilai']

    # Tambahkan data ke Excel
    sheet.append([Nama_karyawan, Nama_kegiatan, Waktu, Nilai])

    # Simpan perubahan
    workbook.save(FILE_NAME)
    
def baca_Jadwal():
    """Membaca data Jadwal dari file Excel dan mengembalikan daftar Jadwal."""
    if not os.path.exists("Data_Jadwal.xlsx"):
        return []  # Jika file tidak ada, kembalikan list kosong

    workbook = openpyxl.load_workbook("Data_Jadwal.xlsx")
    sheet = workbook["Jadwal"]

    jadwal_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        jadwal_list.append(row)  # Menyimpan seluruh baris sebagai tuple

    return jadwal_list

def nilai():
    print("\nMAU MELAKUKAN APA DIniali???")
    print("1. Tambah Nilai")
    print("2. Lihat Nilai")
    print("3. Edit Nilai")
    print("4. Hapus Nilai")
    print("5. Kembali")
    pilihan = input("Pilih Menu: ")
    
    if pilihan == "1":
        tambah()
    elif pilihan == "2":
        baca()  
    elif pilihan == "3":
        edit()
    elif pilihan == "4":
        hapus()
    elif pilihan == "5":
        from main import menu_utama
        menu_utama()
    else:
        print("PILIHAN SALAH!!😒😒")
        
def tambah():
    print("\nISI UNTUK MENAMBAHKAN JADWAL Jadwal!!")

    # Baca daftar Jadwal dari file Jadwal
    Jadwal_list = baca_Jadwal()

    # Tampilkan pilihan Jadwal
    print("Pilih Jadwal:")
    for i, jadwal in enumerate(Jadwal_list, start=1):
        print(f"{i}. {jadwal}")  # Menampilkan seluruh data Jadwal
    print("0. Masukkan Jadwal baru")

    pilihan = int(input("Masukkan pilihan: "))
    if pilihan == 0:
        Nama_karyawan = input("Masukkan Nama Karyawan: ")
    else:
        Nama_karyawan = Jadwal_list[pilihan - 1][0]  
        Nama_kegiatan = Jadwal_list[pilihan - 1][1]  # Ambil email dari tuple
        Waktu = Jadwal_list[pilihan - 1][3]

    # Lanjutkan dengan input lainnya...
    Nilai = input("Masukkan Nilai Pelatihan: ")

    try:
        simpan_ke_excel(Nama_karyawan, Nama_kegiatan, Waktu, Nilai)
        print("Data berhasil ditambahkan ke Nilai.")
        nilai()
    except Exception as e:
        print(f"Terjadi kesalahan saat menyimpan data: {e}")

def baca():
    if not os.path.exists(FILE_NAME):
        print("DATA TIDAK DITEMUKAN")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if "Nilai" not in workbook.sheetnames:  # Check for the correct sheet name
        print("Sheet Belum ada")
        return
    
    sheet = workbook['Nilai']  # Use the correct sheet name
    if sheet.max_row == 1:
        print("Belum ada data")
        return
    
    print("\nDAFTAR KEGIATAN/Nilai")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"\nNama Karyawan: {row[0]}, ")
        print(f"Nama Kegiatan: {row[1]},")
        print(f"Waktu: {row[2]}")
        print(f"Nilai : {row[3]}\n")
    workbook.close()
    nilai()

def edit():
    if not os.path.exists(FILE_NAME):
        print("DATA TIDAK DITEMUKAN")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if "Nilai" not in workbook.sheetnames:
        print("Sheet Belum ada")
        return

    sheet = workbook['Nilai']
    if sheet.max_row == 1:
        print("Belum ada data untuk diedit.")
        return

    # Display current entries
    print("\nDAFTAR KEGIATAN/Nilai UNTUK DIEDIT")
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        print(f"{index}. Nama Karyawan: {row[0]}, Nama Kegiatan: {row[1]}, Waktu: {row[2]}, Nilai: {row[3]}")

    # Ask user for the entry they want to edit
    try:
        pilihan = int(input("Masukkan nomor nilai yang ingin diedit: "))
        if pilihan < 1 or pilihan > sheet.max_row - 1:  # Adjust for header row
            print("Nomor nilai tidak valid.")
            return

        # Get the current data
        current_row = sheet[pilihan + 1]  # +1 because we skip the header row

        # Display current data
        print(f"Data saat ini: Nama Karyawan: {current_row[0].value}, Nama Kegiatan: {current_row[1].value}, Waktu: {current_row[2].value}, Nilai: {current_row[3].value}")
 
        jadwal_list = baca_Jadwal()

        # Tampilkan pilihan nama karyawan
        print("Pilih nama karyawan:")
        for i, Jadwal in enumerate(jadwal_list, start=1):
            print(f"{i}. {Jadwal}")
        print("0. Masukkan nama karyawan baru")

        pilihan_nama = int(input("Masukkan pilihan: "))
        if pilihan_nama == 0:
            new_nama_karyawan = input("Masukkan Nama Karyawan: ")
        else:
            new_nama_karyawan = jadwal_list[pilihan_nama - 1][0]
            new_nama_kegaitan = jadwal_list[pilihan_nama - 1][1]
            new_waktu = jadwal_list[pilihan_nama - 1][3]
            
            new_nilai = input("Masukkan Nilai baru (tekan Enter untuk tetap): ")
            # Update the row with new values
            current_row[0].value = new_nama_karyawan
            current_row[1].value = new_nama_kegaitan
            current_row[2].value = new_waktu
            current_row[3].value = new_nilai

            # Save the changes
            workbook.save(FILE_NAME)
            print("Data berhasil diperbarui.")
    except Exception as e:
        print(f"Terjadi kesalahan saat mengedit data: {e}")
    finally:
        workbook.close()
        nilai()
        
def hapus():
    if not os.path.exists(FILE_NAME):
        print("DATA TIDAK DITEMUKAN")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    if "Nilai" not in workbook.sheetnames:
        print("Sheet Belum ada")
        workbook.close()
        return

    sheet = workbook['Nilai']
    if sheet.max_row == 1:
        print("Belum ada data untuk dihapus.")
        workbook.close()
        return

    # Tampilkan data untuk membantu pengguna memilih
    print("\nDAFTAR KEGIATAN/NILAI UNTUK DIHAPUS")
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        print(f"{index}. Nama Karyawan: {row[0]}, Nama Kegiatan: {row[1]}, Waktu: {row[2]}, Nilai: {row[3]}")

    try:
        # Meminta pengguna untuk memilih data yang ingin dihapus
        pilihan = int(input("Masukkan nomor data yang ingin dihapus: "))
        if pilihan < 1 or pilihan > sheet.max_row - 1:  # Validasi nomor urut
            print("Nomor tidak valid.")
            return

        # Hapus baris dari sheet
        sheet.delete_rows(pilihan + 1)  # +1 karena header ada di baris pertama

        # Simpan perubahan
        workbook.save(FILE_NAME)
        print("Data berhasil dihapus.")
    except ValueError:
        print("Input harus berupa angka.")
    except Exception as e:
        print(f"Terjadi kesalahan saat menghapus data: {e}")
    finally:
        workbook.close()
    nilai()
