import openpyxl
import os

FILE_NAME = "Laporan_Pelatihan.xlsx"

# Fungsi untuk menyimpan data ke Excel
def simpan_ke_excel(Nama_karyawan, Nama_kegiatan, Waktu, Nilai, kehadiran, status_kelulusan):
    # Jika file Excel tidak ada, buat file baru
    if not os.path.exists(FILE_NAME):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Laporan"
        # Header kolom
        sheet.append(['Nama_karyawan', 'Nama_kegiatan', 'Waktu', 'Nilai', 'kehadiran', 'status_kelulusan'])
    else:
        workbook = openpyxl.load_workbook(FILE_NAME)
        sheet = workbook['Laporan']

    # Tambahkan data baru
    sheet.append([Nama_karyawan, Nama_kegiatan, Waktu, Nilai, kehadiran, status_kelulusan])

    # Simpan perubahan
    workbook.save(FILE_NAME)
    
def baca_Nilai():
    """Membaca data Nilai dari file Excel dan mengembalikan daftar Nilai."""
    if not os.path.exists("Data_Nilai.xlsx"):
        return []  # Jika file tidak ada, kembalikan list kosong

    workbook = openpyxl.load_workbook("Data_Nilai.xlsx")
    sheet = workbook["Nilai"]

    Nilai_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        Nilai_list.append(row)  # Menyimpan seluruh baris sebagai tuple

    return Nilai_list
    
def laporan():
    print("\nMAU MELAKUKAN APA DIniali???")
    print("1. Tambah Status")
    print("2. Lihat Satus")
    print("5. Kembali")
    pilihan = input("Pilih Menu: ")
    
    if pilihan == "1":
        masukkan_data()
    elif pilihan == "2":
        lihat_laporan()  
    elif pilihan == "5":
        from main import menu_utama
        menu_utama()
    else:
        print("PILIHAN SALAH!!😒😒")

# Fungsi untuk menentukan status kelulusan
def hitung_status_kehadiran_dan_kelulusan(Kehadiran, Nilai):
    # Pastikan Nilai adalah tipe data numerik
    try:
        Nilai = float(Nilai)  # Mengonversi Nilai ke float
    except ValueError:
        return "Nilai tidak valid"

    # Tentukan status kelulusan berdasarkan kehadiran dan nilai
    if Kehadiran.lower() == "hadir":
        if Nilai > 80:
            return "Lulus"
        else:
            return "Tidak Lulus"
    elif Kehadiran.lower() == "tidak hadir":
        return "Tidak Lulus"
    else:
        return "Input Kehadiran Tidak Valid"

# Fungsi untuk memasukkan data pelatihan
def masukkan_data():
    print("\nISI UNTUK MENAMBAHKAN JADWAL Jadwal!!")

    # Baca daftar Jadwal dari file Jadwal
    Nilai_list = baca_Nilai()

    # Tampilkan pilihan Jadwal
    print("Pilih Nilai:")
    for i, Nilai in enumerate(Nilai_list, start=1):
        print(f"{i}. {Nilai}")  # Menampilkan seluruh data Jadwal
    print("0. Masukkan Nilai baru")

    pilihan = int(input("Masukkan pilihan: "))
    if pilihan == 0:
        Nama_karyawan = input("Masukkan Nama Karyawan: ")
    else:
        Nama_karyawan = Nilai_list[pilihan - 1][0]  
        Nama_kegiatan = Nilai_list[pilihan - 1][1]  # Ambil email dari tuple
        Waktu = Nilai_list[pilihan - 1][2]
        Nilai = Nilai_list [pilihan - 1][3]

    # Lanjutkan dengan input lainnya...
    Kehadiran = input("Masukkan Kehadiran Pelatihan: ")
    status_kelulusan = hitung_status_kehadiran_dan_kelulusan(Kehadiran, Nilai)

    try:
        simpan_ke_excel(Nama_karyawan, Nama_kegiatan, Waktu, Nilai, Kehadiran, status_kelulusan)
        print("Data berhasil ditambahkan ke Nilai.")
        laporan()
    except Exception as e:
        print(f"Terjadi kesalahan saat menyimpan data: {e}")

# Fungsi untuk menampilkan laporan pelatihan
def lihat_laporan():
    # lihat status selesai dan belum selesai
    # lama jadwal pelatihan namabah row tanggal
    #relasi!!!!!!!!!!!!!!!!!
    if not os.path.exists(FILE_NAME):
        print("Laporan tidak ditemukan.")
        return

    workbook = openpyxl.load_workbook(FILE_NAME)
    sheet = workbook['Laporan']
    
    if sheet.max_row == 1:
        print("Belum ada data dalam laporan.")
        workbook.close()
        return

    print("\n=== Laporan Kehadiran dan Hasil Pelatihan ===")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(f"Nama Karyawan: {row[0]}, Nama Kegitan: {row[1]}, Waktu: {row[2]}, Nilai: {row[3]}, Kehadiran: {row[4]}, Status Kelulusan: {row[5]}\n")

    workbook.close()
    laporan()
    


